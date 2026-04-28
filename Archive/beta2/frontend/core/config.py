"""
Frontend config loader.
Reads config.yaml + inventory.xlsx exactly as before.
Also serialises ClusterConfig to JSON dicts for sending to each backend.
"""
from __future__ import annotations
import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

import yaml
import openpyxl


# ── Credential structs ────────────────────────────────────────────────────────

@dataclass
class SSHCred:
    username:    str
    password:    Optional[str] = None
    private_key: Optional[str] = None
    passphrase:  Optional[str] = None
    port:        int = 22

    def __post_init__(self):
        if not self.password and not self.private_key:
            raise ValueError("SSH credential needs 'password' or 'private_key'.")

    def to_dict(self) -> dict:
        return {
            "username":    self.username,
            "password":    self.password,
            "private_key": self.private_key,
            "passphrase":  self.passphrase,
            "port":        self.port,
        }


@dataclass
class APICred:
    token:      Optional[str] = None
    username:   Optional[str] = None
    password:   Optional[str] = None
    verify_ssl: bool = True

    def to_dict(self) -> dict:
        return {
            "token":      self.token,
            "username":   self.username,
            "password":   self.password,
            "verify_ssl": self.verify_ssl,
        }


# ── Per-cluster config ────────────────────────────────────────────────────────

@dataclass
class ClusterConfig:
    name:             str
    cluster_type:     str
    environment:      str = ""
    description:      str = ""
    installer_host:   Optional[str] = None
    api_url:          Optional[str] = None
    nodes:            List[str] = field(default_factory=list)
    ssh_cred:         Optional[SSHCred] = None
    api_cred:         Optional[APICred] = None
    disk_threshold:          Optional[int] = None
    restart_warn_threshold:  Optional[int] = None
    restart_fail_threshold:  Optional[int] = None
    pod_age_min_warn:        Optional[int] = None
    pod_age_min_fail:        Optional[int] = None
    tags:             Dict[str, str] = field(default_factory=dict)
    enabled:          bool = True

    def to_dict(self) -> dict:
        """Serialise for sending to backend via WebSocket."""
        return {
            "name":           self.name,
            "cluster_type":   self.cluster_type,
            "environment":    self.environment,
            "description":    self.description,
            "installer_host": self.installer_host,
            "api_url":        self.api_url,
            "nodes":          self.nodes,
            "ssh_cred":       self.ssh_cred.to_dict() if self.ssh_cred else None,
            "api_cred":       self.api_cred.to_dict() if self.api_cred else None,
            "disk_threshold":         self.disk_threshold,
            "restart_warn_threshold": self.restart_warn_threshold,
            "restart_fail_threshold": self.restart_fail_threshold,
            "pod_age_min_warn":       self.pod_age_min_warn,
            "pod_age_min_fail":       self.pod_age_min_fail,
            "tags":           self.tags,
        }


# ── Global app config ─────────────────────────────────────────────────────────

@dataclass
class AppConfig:
    inventory_path:        Path
    output_dir:            Optional[Path]      = None
    source_bastion_host:   str                 = ""
    source_bastion_port:   int                 = 22
    backend_port:          int                 = 8765
    heartbeat_timeout:     int                 = 60
    max_parallel_clusters: int                 = 5
    max_parallel_nodes:    int                 = 10
    ssh_timeout:           int                 = 30
    cmd_timeout:           int                 = 60
    enabled_checks:        Optional[List[str]] = None
    disk_threshold:         int   = 80
    restart_warn_threshold: int   = 10
    restart_fail_threshold: int   = 50
    pod_age_min_warn:       int   = 5
    pod_age_min_fail:       int   = 2
    cert_warn_days:         int   = 30
    load_ratio_warn:        float = 1.0
    load_ratio_fail:        float = 2.0
    mem_used_pct_warn:      int   = 80
    mem_used_pct_fail:      int   = 90
    swap_used_pct_warn:     int   = 50
    html_report:            bool  = True
    email_friendly:         bool  = True
    clusters:               List[ClusterConfig] = field(default_factory=list)

    def to_backend_dict(self) -> dict:
        """Serialise the config payload sent to each backend."""
        return {
            "max_parallel_clusters": self.max_parallel_clusters,
            "max_parallel_nodes":    self.max_parallel_nodes,
            "ssh_timeout":           self.ssh_timeout,
            "cmd_timeout":           self.cmd_timeout,
            "enabled_checks":        self.enabled_checks,
            "disk_threshold":        self.disk_threshold,
            "restart_warn_threshold":self.restart_warn_threshold,
            "restart_fail_threshold":self.restart_fail_threshold,
            "pod_age_min_warn":      self.pod_age_min_warn,
            "pod_age_min_fail":      self.pod_age_min_fail,
            "cert_warn_days":        self.cert_warn_days,
            "load_ratio_warn":       self.load_ratio_warn,
            "load_ratio_fail":       self.load_ratio_fail,
            "mem_used_pct_warn":     self.mem_used_pct_warn,
            "mem_used_pct_fail":     self.mem_used_pct_fail,
            "swap_used_pct_warn":    self.swap_used_pct_warn,
            "clusters":              [c.to_dict() for c in self.clusters],
        }


# ── Loaders ───────────────────────────────────────────────────────────────────

DEFAULT_CONFIG_PATHS = [
    Path("config/config.yaml"),
    Path("clusterpulse.yaml"),
    Path(os.path.expanduser("~/.clusterpulse/config.yaml")),
]
DEFAULT_INVENTORY_PATHS = [
    Path("config/inventory.xlsx"),
    Path("inventory.xlsx"),
]


def _find_file(explicit: Optional[str], defaults: List[Path], label: str) -> Path:
    if explicit:
        p = Path(explicit)
        if not p.exists():
            raise FileNotFoundError(f"{label} not found: {p}")
        return p
    for p in defaults:
        if p.exists():
            return p
    raise FileNotFoundError(f"{label} not found. Tried: {[str(d) for d in defaults]}")


def load_app_config(
    config_path:    Optional[str] = None,
    inventory_path: Optional[str] = None,
    output_dir:     Optional[str] = None,
    cluster_type:   Optional[str] = None,
    checks:         Optional[str] = None,
    max_parallel:   Optional[int] = None,
    ssh_timeout:    Optional[int] = None,
) -> AppConfig:
    cfg_path = _find_file(config_path, DEFAULT_CONFIG_PATHS, "config.yaml")
    with open(cfg_path) as f:
        raw = yaml.safe_load(f) or {}

    inv_path_str = inventory_path or raw.get("inventory_path")
    inv_path     = _find_file(inv_path_str, DEFAULT_INVENTORY_PATHS, "inventory.xlsx")
    out_dir      = Path(output_dir) if output_dir else Path(raw.get("output_dir", "results"))

    thresholds  = raw.get("thresholds",  {})
    parallelism = raw.get("parallelism", {})
    reports     = raw.get("reports",     {})
    source      = raw.get("source_bastion", {})

    app = AppConfig(
        inventory_path        = inv_path,
        output_dir            = out_dir,
        source_bastion_host   = source.get("host", ""),
        source_bastion_port   = int(source.get("port", 22)),
        backend_port          = int(raw.get("backend_port", 8765)),
        heartbeat_timeout     = int(raw.get("heartbeat_timeout", 60)),
        max_parallel_clusters = max_parallel or parallelism.get("max_parallel_clusters", 5),
        max_parallel_nodes    = parallelism.get("max_parallel_nodes", 10),
        ssh_timeout           = ssh_timeout or parallelism.get("ssh_timeout", 30),
        cmd_timeout           = parallelism.get("cmd_timeout", 60),
        enabled_checks        = [c.strip() for c in checks.split(",")] if checks else raw.get("enabled_checks"),
        disk_threshold        = thresholds.get("disk_pct",           80),
        restart_warn_threshold= thresholds.get("restart_warn",       10),
        restart_fail_threshold= thresholds.get("restart_fail",       50),
        pod_age_min_warn      = thresholds.get("pod_age_min_warn_m",  5),
        pod_age_min_fail      = thresholds.get("pod_age_min_fail_m",  2),
        cert_warn_days        = thresholds.get("cert_warn_days",     30),
        load_ratio_warn       = thresholds.get("load_ratio_warn",   1.0),
        load_ratio_fail       = thresholds.get("load_ratio_fail",   2.0),
        mem_used_pct_warn     = thresholds.get("mem_used_pct_warn",  80),
        mem_used_pct_fail     = thresholds.get("mem_used_pct_fail",  90),
        swap_used_pct_warn    = thresholds.get("swap_used_pct_warn", 50),
        html_report           = reports.get("html",           True),
        email_friendly        = reports.get("email_friendly", True),
    )
    app.clusters = _load_inventory(inv_path, cluster_type)
    return app


def _load_inventory(path: Path, cluster_type_filter: Optional[str]) -> List[ClusterConfig]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Clusters" not in wb.sheetnames:
        raise ValueError("inventory.xlsx must have a 'Clusters' sheet.")

    ws = wb["Clusters"]
    header_row = 3
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5), start=1):
        first = str(row[0].value).strip().lower() if row[0].value else ""
        if first == "cluster_name":
            header_row = i
            break
    headers    = [str(c.value).strip().lower() if c.value else ""
                  for c in next(ws.iter_rows(min_row=header_row, max_row=header_row))]
    data_start = header_row + 1

    def col(row, name, default=None):
        try:
            idx = headers.index(name)
            v   = row[idx].value
            return str(v).strip() if v is not None else default
        except (ValueError, IndexError):
            return default

    def bool_col(row, name, default=True):
        v = col(row, name, str(default))
        return str(v).lower() not in ("false", "no", "0", "disabled")

    def int_or_none(v):
        return int(v) if v and str(v).isdigit() else None

    clusters: Dict[str, ClusterConfig] = {}
    for row in ws.iter_rows(min_row=data_start):
        if all(c.value is None for c in row): continue
        name = col(row, "cluster_name")
        if not name or name.lower() == "none": continue
        ctype = (col(row, "type") or "ocp").lower()
        if cluster_type_filter and ctype != cluster_type_filter.lower(): continue
        if not bool_col(row, "enabled", True): continue

        ssh_user = col(row, "ssh_username")
        ssh_cred = None
        if ssh_user:
            try:
                ssh_cred = SSHCred(
                    username    = ssh_user,
                    password    = col(row, "ssh_password"),
                    private_key = col(row, "ssh_private_key"),
                    port        = int(col(row, "ssh_port") or 22),
                )
            except ValueError:
                pass

        api_token = col(row, "api_token")
        api_user  = col(row, "api_username")
        api_cred  = APICred(
            token      = api_token,
            username   = api_user,
            password   = col(row, "api_password"),
            verify_ssl = bool_col(row, "verify_ssl", True),
        ) if (api_token or api_user) else None

        clusters[name] = ClusterConfig(
            name           = name,
            cluster_type   = ctype,
            environment    = col(row, "environment") or "",
            description    = col(row, "description") or "",
            installer_host = col(row, "installer_host"),
            api_url        = col(row, "api_url"),
            ssh_cred       = ssh_cred,
            api_cred       = api_cred,
            disk_threshold         = int_or_none(col(row, "disk_threshold")),
            restart_warn_threshold = int_or_none(col(row, "restart_warn")),
            restart_fail_threshold = int_or_none(col(row, "restart_fail")),
            pod_age_min_warn       = int_or_none(col(row, "pod_age_warn_m")),
            pod_age_min_fail       = int_or_none(col(row, "pod_age_fail_m")),
        )

    # Nodes sheet
    if "Nodes" in wb.sheetnames:
        wn = wb["Nodes"]
        nh = [str(c.value).strip().lower() if c.value else ""
              for c in next(wn.iter_rows(min_row=1, max_row=1))]

        def ncol(row, name, default=None):
            try:
                idx = nh.index(name)
                v   = row[idx].value
                return str(v).strip() if v is not None else default
            except (ValueError, IndexError):
                return default

        for row in wn.iter_rows(min_row=2):
            if all(c.value is None for c in row): continue
            cname    = ncol(row, "cluster_name")
            node_ip  = ncol(row, "node_ip") or ncol(row, "hostname")
            if not cname or not node_ip: continue
            if cname in clusters:
                clusters[cname].nodes.append(node_ip)

    return list(clusters.values())
