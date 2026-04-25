"""
Config loader for ClusterPulse.
- config.yaml  : all tunable parameters, thresholds, paths
- inventory.xlsx: cluster/node rows (two sheets: Clusters, Nodes)
"""
from __future__ import annotations
import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

import yaml
import openpyxl


# ─────────────────────────── credential structs ──────────────────────────────

@dataclass
class SSHCred:
    username:   str
    password:   Optional[str] = None
    private_key: Optional[str] = None
    passphrase: Optional[str] = None
    port:       int = 22

    def __post_init__(self):
        if not self.password and not self.private_key:
            raise ValueError("SSH credential needs 'password' or 'private_key'.")


@dataclass
class APICred:
    token:      Optional[str] = None
    username:   Optional[str] = None
    password:   Optional[str] = None
    verify_ssl: bool = True


# ─────────────────────────── per-cluster config ───────────────────────────────

@dataclass
class ClusterConfig:
    name:             str
    cluster_type:     str           # "ocp" | "cvim"
    environment:      str = ""
    description:      str = ""
    installer_host:   Optional[str] = None
    api_url:          Optional[str] = None
    nodes:            List[str] = field(default_factory=list)
    ssh_cred:         Optional[SSHCred] = None
    api_cred:         Optional[APICred] = None
    # threshold overrides (None → use global from AppConfig)
    disk_threshold:          Optional[int]  = None
    restart_warn_threshold:  Optional[int]  = None
    restart_fail_threshold:  Optional[int]  = None
    pod_age_min_warn:        Optional[int]  = None
    pod_age_min_fail:        Optional[int]  = None
    tags:             Dict[str, str] = field(default_factory=dict)
    enabled:          bool = True


# ─────────────────────────── global app config ───────────────────────────────

@dataclass
class AppConfig:
    # paths
    inventory_path: Path
    output_dir:     Optional[Path]  = None
    log_dir:        Optional[Path]  = None

    # parallelism
    max_parallel_clusters: int = 5
    max_parallel_nodes:    int = 10
    ssh_timeout:           int = 30
    cmd_timeout:           int = 60

    # check filters (None = run all)
    enabled_checks: Optional[List[str]] = None

    # global thresholds
    disk_threshold:         int = 80
    restart_warn_threshold: int = 10
    restart_fail_threshold: int = 50
    pod_age_min_warn:       int = 5
    pod_age_min_fail:       int = 2
    cert_warn_days:         int = 30
    load_ratio_warn:        float = 1.0
    load_ratio_fail:        float = 2.0
    mem_used_pct_warn:      int = 80
    mem_used_pct_fail:      int = 90
    swap_used_pct_warn:     int = 50

    # report options
    html_report:     bool = True
    email_friendly:  bool = True    # inline styles for email clients
    verbose_console: bool = False

    # clusters (populated by inventory loader)
    clusters: List[ClusterConfig] = field(default_factory=list)


# ─────────────────────────── loaders ─────────────────────────────────────────

DEFAULT_CONFIG_PATHS = [
    Path("config/config.yaml"),
    Path("clusterpulse.yaml"),
    Path(os.path.expanduser("~/.clusterpulse/config.yaml")),
]

DEFAULT_INVENTORY_PATHS = [
    Path("config/inventory.xlsx"),
    Path("inventory.xlsx"),
]


def _find_file(explicit: Optional[str], defaults: List[Path], label: str) -> Path:
    if explicit:
        p = Path(explicit)
        if not p.exists():
            raise FileNotFoundError(f"{label} not found: {p}")
        return p
    for p in defaults:
        if p.exists():
            return p
    raise FileNotFoundError(
        f"{label} not found. Tried: {[str(d) for d in defaults]}. "
        f"Pass it explicitly with the appropriate CLI flag."
    )


def load_app_config(
    config_path:    Optional[str] = None,
    inventory_path: Optional[str] = None,
    output_dir:     Optional[str] = None,
    cluster_type:   Optional[str] = None,
    checks:         Optional[str] = None,
    max_parallel:   Optional[int] = None,
    ssh_timeout:    Optional[int] = None,
    verbose:        bool = False,
) -> AppConfig:
    """Load config.yaml then overlay CLI overrides, then load inventory.xlsx."""

    cfg_path = _find_file(config_path, DEFAULT_CONFIG_PATHS, "config.yaml")
    with open(cfg_path) as f:
        raw = yaml.safe_load(f) or {}

    inv_path_str = inventory_path or raw.get("inventory_path")
    inv_path = _find_file(inv_path_str, DEFAULT_INVENTORY_PATHS, "inventory.xlsx")

    out_dir = (
        Path(output_dir) if output_dir
        else Path(raw.get("output_dir", "results"))
    )
    log_dir = Path(raw.get("log_dir", str(out_dir)))

    thresholds = raw.get("thresholds", {})
    parallelism = raw.get("parallelism", {})
    reports = raw.get("reports", {})

    app = AppConfig(
        inventory_path = inv_path,
        output_dir     = out_dir,
        log_dir        = log_dir,
        max_parallel_clusters = max_parallel or parallelism.get("max_parallel_clusters", 5),
        max_parallel_nodes    = parallelism.get("max_parallel_nodes", 10),
        ssh_timeout           = ssh_timeout or parallelism.get("ssh_timeout", 30),
        cmd_timeout           = parallelism.get("cmd_timeout", 60),
        enabled_checks        = [c.strip() for c in checks.split(",")] if checks else raw.get("enabled_checks"),
        disk_threshold        = thresholds.get("disk_pct",           80),
        restart_warn_threshold= thresholds.get("restart_warn",       10),
        restart_fail_threshold= thresholds.get("restart_fail",       50),
        pod_age_min_warn      = thresholds.get("pod_age_min_warn_m",  5),
        pod_age_min_fail      = thresholds.get("pod_age_min_fail_m",  2),
        cert_warn_days        = thresholds.get("cert_warn_days",     30),
        load_ratio_warn       = thresholds.get("load_ratio_warn",   1.0),
        load_ratio_fail       = thresholds.get("load_ratio_fail",   2.0),
        mem_used_pct_warn     = thresholds.get("mem_used_pct_warn",  80),
        mem_used_pct_fail     = thresholds.get("mem_used_pct_fail",  90),
        swap_used_pct_warn    = thresholds.get("swap_used_pct_warn", 50),
        html_report           = reports.get("html",          True),
        email_friendly        = reports.get("email_friendly", True),
        verbose_console       = verbose or raw.get("verbose", False),
    )

    # Load inventory and filter
    app.clusters = _load_inventory(inv_path, app, cluster_type)
    return app


def _load_inventory(path: Path, app: AppConfig, cluster_type_filter: Optional[str]) -> List[ClusterConfig]:
    wb = openpyxl.load_workbook(path, data_only=True)

    # ── Sheet 1: Clusters ────────────────────────────────────────────────────
    clusters_by_name: Dict[str, ClusterConfig] = {}

    if "Clusters" not in wb.sheetnames:
        raise ValueError("inventory.xlsx must have a 'Clusters' sheet.")

    ws = wb["Clusters"]
    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def col(row, name: str, default=None):
        try:
            idx = headers.index(name)
            v = row[idx].value
            return str(v).strip() if v is not None else default
        except (ValueError, IndexError):
            return default

    def bool_col(row, name: str, default: bool = True) -> bool:
        v = col(row, name, str(default))
        return str(v).lower() not in ("false", "no", "0", "disabled")

    for row in ws.iter_rows(min_row=2):
        if all(c.value is None for c in row):
            continue
        name = col(row, "cluster_name")
        if not name or name.lower() == "none":
            continue

        ctype = (col(row, "type") or "ocp").lower()
        if cluster_type_filter and ctype != cluster_type_filter.lower():
            continue
        if not bool_col(row, "enabled", True):
            continue

        # SSH credential
        ssh_user = col(row, "ssh_username")
        ssh_pass = col(row, "ssh_password")
        ssh_key  = col(row, "ssh_private_key")
        ssh_port = int(col(row, "ssh_port") or 22)
        ssh_cred = None
        if ssh_user:
            try:
                ssh_cred = SSHCred(
                    username=ssh_user, password=ssh_pass,
                    private_key=ssh_key, port=ssh_port
                )
            except ValueError:
                pass

        # API credential
        api_token = col(row, "api_token")
        api_user  = col(row, "api_username")
        api_pass  = col(row, "api_password")
        verify    = bool_col(row, "verify_ssl", True)
        api_cred  = APICred(token=api_token, username=api_user,
                            password=api_pass, verify_ssl=verify) if (api_token or api_user) else None

        # Threshold overrides (blank → None → use global)
        def int_or_none(v): return int(v) if v and str(v).isdigit() else None

        cc = ClusterConfig(
            name           = name,
            cluster_type   = ctype,
            environment    = col(row, "environment") or "",
            description    = col(row, "description") or "",
            installer_host = col(row, "installer_host"),
            api_url        = col(row, "api_url"),
            ssh_cred       = ssh_cred,
            api_cred       = api_cred,
            disk_threshold         = int_or_none(col(row, "disk_threshold")),
            restart_warn_threshold = int_or_none(col(row, "restart_warn")),
            restart_fail_threshold = int_or_none(col(row, "restart_fail")),
            pod_age_min_warn       = int_or_none(col(row, "pod_age_warn_m")),
            pod_age_min_fail       = int_or_none(col(row, "pod_age_fail_m")),
            enabled        = True,
        )
        clusters_by_name[name] = cc

    # ── Sheet 2: Nodes (optional) ────────────────────────────────────────────
    if "Nodes" in wb.sheetnames:
        wn = wb["Nodes"]
        n_headers = [str(c.value).strip().lower() if c.value else "" for c in next(wn.iter_rows(min_row=1, max_row=1))]

        def ncol(row, name, default=None):
            try:
                idx = n_headers.index(name)
                v = row[idx].value
                return str(v).strip() if v is not None else default
            except (ValueError, IndexError):
                return default

        for row in wn.iter_rows(min_row=2):
            if all(c.value is None for c in row):
                continue
            cluster_name = ncol(row, "cluster_name")
            node_ip      = ncol(row, "node_ip") or ncol(row, "hostname")
            if not cluster_name or not node_ip:
                continue
            if cluster_name in clusters_by_name:
                # Allow per-node SSH credential override
                n_user = ncol(row, "ssh_username")
                n_pass = ncol(row, "ssh_password")
                n_key  = ncol(row, "ssh_private_key")
                n_port = int(ncol(row, "ssh_port") or 22)
                if n_user and not clusters_by_name[cluster_name].ssh_cred:
                    try:
                        clusters_by_name[cluster_name].ssh_cred = SSHCred(
                            username=n_user, password=n_pass, private_key=n_key, port=n_port)
                    except ValueError:
                        pass
                clusters_by_name[cluster_name].nodes.append(node_ip)

    return list(clusters_by_name.values())


def resolve_threshold(cluster_val: Optional[int], global_val: int) -> int:
    return cluster_val if cluster_val is not None else global_val
