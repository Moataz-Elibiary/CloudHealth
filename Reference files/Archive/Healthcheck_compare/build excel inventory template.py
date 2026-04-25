from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.numbers import FORMAT_TEXT

wb = Workbook()

# ── Colour palette ────────────────────────────────────────────────────────────
HDR_BG    = "0F1117"   # near-black header
HDR_FG    = "F59E0B"   # amber text
SEC_BG    = "1E2333"   # section sub-header
SEC_FG    = "60A5FA"   # blue text
OCP_BG    = "0D2137"   # OCP row tint
CVIM_BG   = "0D1F1A"   # CVIM row tint
ALT_BG    = "171923"
OPT_BG    = "1A1E2A"   # optional fields
PASS_BG   = "14291F"
BORDER_C  = "2A3350"
AMBER_BG  = "2D2010"

def hdr_cell(ws, cell_ref, val, fg=HDR_FG, bg=HDR_BG, bold=True, size=10, wrap=True, align="center"):
    c = ws[cell_ref]
    c.value = val
    c.font = Font(name="Consolas", bold=bold, color=fg, size=size)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return c

def data_cell(ws, cell_ref, val, fg="C8D0E0", bg=ALT_BG, italic=False, mono=False, align="left"):
    c = ws[cell_ref]
    c.value = val
    font_name = "Consolas" if mono else "Calibri"
    c.font = Font(name=font_name, color=fg, size=9, italic=italic)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False, shrink_to_fit=False)
    return c

def thin_border():
    s = Side(style="thin", color=BORDER_C)
    return Border(left=s, right=s, top=s, bottom=s)

def apply_border(ws, min_col, max_col, row):
    for col in range(min_col, max_col+1):
        ws.cell(row=row, column=col).border = thin_border()

# ══════════════════════════════════════════════════════════════════════════════
#  Sheet 1 — Clusters
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Clusters"
ws1.sheet_view.showGridLines = False
ws1.sheet_properties.tabColor = "F59E0B"

# Row 1 — sheet title
ws1.merge_cells("A1:W1")
c = ws1["A1"]
c.value = "⚡  ClusterPulse — Cluster Inventory"
c.font = Font(name="Consolas", bold=True, size=14, color=HDR_FG)
c.fill = PatternFill("solid", fgColor=HDR_BG)
c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
ws1.row_dimensions[1].height = 32

# Row 2 — section labels
SECTIONS = [
    ("A2", "CLUSTER IDENTITY",     "A2:D2",  HDR_BG,  HDR_FG),
    ("E2", "SSH CREDENTIALS",      "E2:I2",  "1A1A2E","A78BFA"),
    ("J2", "API CREDENTIALS",      "J2:M2",  "1A2A1A","34D399"),
    ("N2", "THRESHOLDS (OPTIONAL)","N2:S2",  "2A1A10","F59E0B"),
    ("T2", "NODE DISCOVERY",       "T2:T2",  "1A2A2A","67E8F9"),
    ("U2", "TAGS / META",          "U2:W2",  "2A1A2A","F9A8D4"),
]
for cell, label, merge, bg, fg in SECTIONS:
    ws1.merge_cells(merge)
    c = ws1[cell]
    c.value = label
    c.font = Font(name="Consolas", bold=True, size=8, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 18

# Row 3 — column headers
COLS = [
    # (col_letter, header, width, bg, fg, comment)
    ("A", "cluster_name",    22, HDR_BG, HDR_FG,  "Unique cluster name — used in report headings"),
    ("B", "type",            10, HDR_BG, HDR_FG,  "ocp  or  cvim"),
    ("C", "environment",     14, HDR_BG, "94A3B8","e.g. production / staging / dev"),
    ("D", "enabled",         10, HDR_BG, "94A3B8","TRUE or FALSE — set FALSE to skip"),
    ("E", "installer_host",  20, "1A1A2E","A78BFA","IP/hostname of bastion or CVIM installer node"),
    ("F", "ssh_username",    14, "1A1A2E","A78BFA","SSH username (e.g. root / core)"),
    ("G", "ssh_password",    18, "1A1A2E","A78BFA","SSH password (leave blank if using private key)"),
    ("H", "ssh_private_key", 26, "1A1A2E","A78BFA","Path to private key e.g. /home/user/.ssh/id_rsa"),
    ("I", "ssh_port",        10, "1A1A2E","A78BFA","SSH port (default 22)"),
    ("J", "api_url",         34, "1A2A1A","34D399","OCP API URL e.g. https://api.cluster.example.com:6443"),
    ("K", "api_token",       38, "1A2A1A","34D399","OCP service account token (sha256~...)"),
    ("L", "api_username",    16, "1A2A1A","34D399","API username (if token not used)"),
    ("M", "verify_ssl",      12, "1A2A1A","34D399","TRUE or FALSE — verify SSL cert"),
    ("N", "disk_threshold",  14, "2A1A10","F59E0B","Disk % FAIL threshold (blank=use config.yaml)"),
    ("O", "restart_warn",    14, "2A1A10","F59E0B","Restart count WARN threshold"),
    ("P", "restart_fail",    14, "2A1A10","F59E0B","Restart count FAIL threshold"),
    ("Q", "pod_age_warn_m",  15, "2A1A10","F59E0B","Pod age WARN minutes"),
    ("R", "pod_age_fail_m",  15, "2A1A10","F59E0B","Pod age FAIL minutes"),
    ("S", "cert_warn_days",  15, "2A1A10","F59E0B","Cert expiry WARN days"),
    ("T", "description",     28, "1A2A2A","67E8F9","Free-text description of this cluster"),
    ("U", "tag_env",         14, "2A1A2A","F9A8D4","Tag: environment"),
    ("V", "tag_site",        14, "2A1A2A","F9A8D4","Tag: site / datacenter"),
    ("W", "tag_team",        14, "2A1A2A","F9A8D4","Tag: team / owner"),
]

for col_letter, header, width, bg, fg, comment in COLS:
    c = ws1[f"{col_letter}3"]
    c.value = header
    c.font = Font(name="Consolas", bold=True, size=8, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()
    ws1.column_dimensions[col_letter].width = width
ws1.row_dimensions[3].height = 30

# ── Data validation — type column ─────────────────────────────────────────────
dv_type = DataValidation(type="list", formula1='"ocp,cvim"', showErrorMessage=True)
ws1.add_data_validation(dv_type)
dv_type.sqref = "B4:B200"

dv_bool = DataValidation(type="list", formula1='"TRUE,FALSE"')
ws1.add_data_validation(dv_bool)
dv_bool.sqref = "D4:D200"

dv_ssl = DataValidation(type="list", formula1='"TRUE,FALSE"')
ws1.add_data_validation(dv_ssl)
dv_ssl.sqref = "M4:M200"

# ── Sample data rows ──────────────────────────────────────────────────────────
sample_rows = [
    # OCP prod cluster
    ["prod-ocp-dc1",   "ocp",  "production", "TRUE",
     "192.168.10.5",   "root",  "",           "/home/user/.ssh/id_rsa", "22",
     "https://api.prod-ocp-dc1.example.com:6443",
     "sha256~xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx",
     "", "FALSE",
     "85","15","60","5","2","30",
     "Production OCP cluster in DC1",
     "production","DC1","platform-team"],
    # OCP staging
    ["staging-ocp-dc2","ocp",  "staging",    "TRUE",
     "10.0.1.50",      "admin","StgPass456!", "",                        "22",
     "https://api.staging-ocp.example.com:6443",
     "sha256~stagingtoken000000000000000000000000000000000",
     "", "FALSE",
     "","","","","","",
     "Staging OCP cluster",
     "staging","DC2","platform-team"],
    # CVIM prod site A
    ["prod-cvim-siteA","cvim", "production", "TRUE",
     "172.16.0.10",    "root", "CvimRootPw!", "",                       "22",
     "","","","",
     "75","","","","","",
     "Production CVIM OpenStack — Site A",
     "production","Site-A","nfv-team"],
    # CVIM prod site B — disabled
    ["prod-cvim-siteB","cvim", "production", "FALSE",
     "172.16.50.10",   "root", "AnotherPw!", "",                       "22",
     "","","","",
     "","","","","","",
     "Site B — temporarily disabled",
     "production","Site-B","nfv-team"],
]

BG_MAP = {"ocp": OCP_BG, "cvim": CVIM_BG}
for row_idx, row_data in enumerate(sample_rows, start=4):
    bg = BG_MAP.get(row_data[1], ALT_BG)
    row_bg = bg if row_idx % 2 == 0 else ALT_BG
    is_disabled = row_data[3].upper() == "FALSE"
    fg_color = "555577" if is_disabled else "C8D0E0"
    for col_idx, val in enumerate(row_data, start=1):
        col_letter = get_column_letter(col_idx)
        c = ws1.cell(row=row_idx, column=col_idx, value=val)
        c.font = Font(name="Consolas", size=9, color=fg_color,
                      italic=is_disabled,
                      bold=(col_idx == 1))
        c.fill = PatternFill("solid", fgColor=row_bg)
        c.alignment = Alignment(vertical="center", horizontal="left" if col_idx==1 else "center")
        c.border = thin_border()
    ws1.row_dimensions[row_idx].height = 20

# Freeze panes
ws1.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════════════════
#  Sheet 2 — Nodes
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Nodes")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.tabColor = "60A5FA"

# Title
ws2.merge_cells("A1:L1")
c = ws2["A1"]
c.value = "⚡  ClusterPulse — Node Inventory  (optional — nodes auto-discovered if omitted)"
c.font = Font(name="Consolas", bold=True, size=12, color="60A5FA")
c.fill = PatternFill("solid", fgColor=HDR_BG)
c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
ws2.row_dimensions[1].height = 28

NODE_COLS = [
    ("A", "cluster_name",    22, "Cluster name — must match exactly the cluster_name in Clusters sheet"),
    ("B", "node_ip",         20, "Node IP address or hostname"),
    ("C", "hostname",        24, "Optional: node hostname (for display)"),
    ("D", "node_type",       14, "control / compute / storage / worker / master / infra"),
    ("E", "ssh_username",    14, "SSH username (overrides cluster default)"),
    ("F", "ssh_password",    18, "SSH password (overrides cluster default)"),
    ("G", "ssh_private_key", 28, "SSH private key path (overrides cluster default)"),
    ("H", "ssh_port",        10, "SSH port (overrides cluster default)"),
    ("I", "enabled",         10, "TRUE or FALSE"),
    ("J", "notes",           30, "Free-text notes about this node"),
    ("K", "rack",            14, "Optional: rack/cabinet location"),
    ("L", "hardware_model",  20, "Optional: hardware model e.g. UCS C220 M6"),
]

for col_letter, header, width, comment in NODE_COLS:
    c = ws2[f"{col_letter}2"]
    c.value = header
    c.font = Font(name="Consolas", bold=True, size=8, color="60A5FA")
    c.fill = PatternFill("solid", fgColor="1A1A30")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()
    ws2.column_dimensions[col_letter].width = width
ws2.row_dimensions[2].height = 28

# Sample nodes
node_rows = [
    ["prod-ocp-dc1",   "192.168.10.100","master-01","master",   "core","","",  "22","TRUE","Control plane 1","Rack-A01","UCS C220 M6"],
    ["prod-ocp-dc1",   "192.168.10.101","master-02","master",   "core","","",  "22","TRUE","Control plane 2","Rack-A01","UCS C220 M6"],
    ["prod-ocp-dc1",   "192.168.10.102","master-03","master",   "core","","",  "22","TRUE","Control plane 3","Rack-A02","UCS C220 M6"],
    ["prod-ocp-dc1",   "192.168.10.110","worker-01","worker",   "core","","",  "22","TRUE","Compute worker 1","Rack-B01","UCS C240 M6"],
    ["prod-ocp-dc1",   "192.168.10.111","worker-02","worker",   "core","","",  "22","TRUE","Compute worker 2","Rack-B01","UCS C240 M6"],
    ["prod-cvim-siteA","172.16.0.20",   "ctrl-01",  "control",  "root","CvimPw","","22","TRUE","Controller 1","Rack-C01","UCS C220 M5"],
    ["prod-cvim-siteA","172.16.0.21",   "ctrl-02",  "control",  "root","CvimPw","","22","TRUE","Controller 2","Rack-C01","UCS C220 M5"],
    ["prod-cvim-siteA","172.16.1.10",   "comp-01",  "compute",  "root","CvimPw","","22","TRUE","Compute 1",   "Rack-D01","UCS C240 M5"],
    ["prod-cvim-siteA","172.16.1.11",   "comp-02",  "compute",  "root","CvimPw","","22","TRUE","Compute 2",   "Rack-D01","UCS C240 M5"],
    ["prod-cvim-siteA","172.16.2.10",   "stor-01",  "storage",  "root","CvimPw","","22","TRUE","Storage 1",   "Rack-E01","UCS C260 M4"],
]
for row_idx, row_data in enumerate(node_rows, start=3):
    bg = OCP_BG if "ocp" in row_data[0].lower() else CVIM_BG
    for col_idx, val in enumerate(row_data, start=1):
        c = ws2.cell(row=row_idx, column=col_idx, value=val)
        c.font = Font(name="Consolas", size=9, color="C8D0E0",
                      bold=(col_idx <= 2))
        c.fill = PatternFill("solid", fgColor=bg if row_idx%2==0 else ALT_BG)
        c.alignment = Alignment(vertical="center", horizontal="left" if col_idx<=3 else "center")
        c.border = thin_border()
    ws2.row_dimensions[row_idx].height = 18

ws2.freeze_panes = "A3"

# ══════════════════════════════════════════════════════════════════════════════
#  Sheet 3 — Instructions
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Instructions")
ws3.sheet_view.showGridLines = False
ws3.sheet_properties.tabColor = "22C55E"

ws3.merge_cells("A1:E1")
c = ws3["A1"]
c.value = "⚡  ClusterPulse — Quick Reference"
c.font = Font(name="Consolas", bold=True, size=14, color=HDR_FG)
c.fill = PatternFill("solid", fgColor=HDR_BG)
c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
ws3.row_dimensions[1].height = 30

instructions = [
    ("", ""),
    ("GETTING STARTED", ""),
    ("1. Fill the Clusters sheet", "One row per cluster. Required: cluster_name, type, installer_host, ssh credentials"),
    ("2. Fill the Nodes sheet (optional)", "Add node IPs for direct SSH checks. If omitted, nodes are auto-discovered."),
    ("3. Edit config/config.yaml", "Set thresholds, parallelism, output paths."),
    ("4. Run the tool", "python clusterpulse.py  (uses config/inventory.xlsx automatically)"),
    ("", ""),
    ("CLI EXAMPLES", ""),
    ("python clusterpulse.py", "Use default config/config.yaml and config/inventory.xlsx"),
    ("python clusterpulse.py -i my_inv.xlsx -c my_cfg.yaml", "Specify custom paths"),
    ("python clusterpulse.py --type ocp", "OCP clusters only"),
    ("python clusterpulse.py --type cvim --parallel 8", "CVIM clusters, 8 parallel"),
    ("python clusterpulse.py --checks nodes,pods,ceph,host", "Run only specific check categories"),
    ("python clusterpulse.py -o /reports/2024-01", "Custom output directory"),
    ("python clusterpulse.py --no-email --verbose", "Skip email report, verbose console"),
    ("", ""),
    ("COLUMN NOTES", ""),
    ("cluster_name", "Must be unique. Used as report heading and output file name."),
    ("type",         "ocp  or  cvim  (case-insensitive)"),
    ("enabled",      "Set to FALSE to skip a cluster without deleting the row."),
    ("installer_host","For OCP: bastion where 'oc' CLI is available and logged in."),
    ("             ","For CVIM: management/installer node where ciscovim and openrc are configured."),
    ("ssh_password", "Leave blank if using private key. Either password or private_key is required."),
    ("api_token",    "OCP service account token. Get it with: oc serviceaccounts get-token <name> -n <ns>"),
    ("verify_ssl",   "Set FALSE for self-signed OCP certificates (common in lab environments)."),
    ("Thresholds",   "All threshold columns are optional — blank means use config.yaml global value."),
    ("", ""),
    ("OCP CHECK CATEGORIES", ""),
    ("version",      "ClusterVersion conditions: Available, Progressing, Degraded"),
    ("operators",    "All ClusterOperators healthy"),
    ("nodes",        "Node Ready state + role counts"),
    ("pressure",     "MemoryPressure / DiskPressure / PIDPressure per node"),
    ("node_disk",    "Disk utilization via oc debug (first 3 nodes)"),
    ("etcd",         "etcd pods + etcdctl endpoint health + backup"),
    ("controlplane", "apiserver / controller-manager / scheduler pods"),
    ("ceph",         "ODF/Ceph pods, cluster health, OSDs, PGs, StorageCluster CR"),
    ("pvcs",         "PVCs not Lost/Pending + PV states"),
    ("storageclasses","Default StorageClass present"),
    ("pods",         "All-namespace pod audit: status, restarts, age"),
    ("deployments",  "Deployments + StatefulSets fully available"),
    ("daemonsets",   "DaemonSets with full node coverage"),
    ("jobs",         "Failed Jobs and CronJob suspended state"),
    ("hpa",          "HPAs at maximum replicas (capacity pressure)"),
    ("network",      "CNI operator, DNS pods, OVN/SDN"),
    ("ingress",      "IngressControllers + router pod health"),
    ("events",       "Warning events cluster-wide"),
    ("certs",        "TLS certificate expiry scan"),
    ("mcp",          "MachineConfigPools not degraded/updating"),
    ("nodes_upgrade","Node OS version consistency + cordoned nodes"),
    ("quota",        "ResourceQuotas and LimitRanges"),
    ("rbac",         "Privileged SCC usage + cluster-admin bindings"),
    ("alerts",       "Prometheus critical/warning firing alerts"),
    ("logging",      "Cluster Logging / Loki stack pods"),
    ("imageregistry","Image registry operator availability"),
    ("backup",       "ETCD backup resource freshness"),
    ("", ""),
    ("CVIM CHECK CATEGORIES", ""),
    ("hypervisors",  "Hypervisor count UP vs configured, per-HV state, CPU/RAM stats"),
    ("network",      "Neutron agent count and liveness"),
    ("volumes",      "Cinder volume services"),
    ("compute_svc",  "Nova compute service state"),
    ("identity",     "Keystone token issue + endpoint list"),
    ("image_svc",    "Glance image count and state"),
    ("cloudpulse",   "Cloudpulse test results"),
    ("vms",          "VM count by status"),
    ("vm_errors",    "VMs in ERROR or unusual task states"),
    ("rabbitmq",     "Functional tests via rabbit_api.py or rabbitmqctl cluster_status"),
    ("mariadb",      "Galera wsrep cluster size, state, readiness"),
    ("memcached",    "Memcached uptime via stats"),
    ("containers",   "Podman/Docker container running vs desired per node"),
    ("ceph",         "Ceph cluster health, OSD and I/O summary"),
    ("ceph_pools",   "Ceph pool list, PG status, OSD tree"),
    ("ovs",          "OVS bridge inventory and version"),
    ("haproxy",      "HAProxy config check + VIP detection + stats"),
    ("nfs",          "NFS exports and active mounts"),
    ("installer",    "CVIM version, management health, recent error logs"),
    ("", ""),
    ("HOST CHECK CATEGORIES (both cluster types)", ""),
    ("host",         "Uptime/load, OS info, CPU model/governor/throttle, "
                     "RAM usage/OOM/NUMA, Disk/SMART/inode/IOstat, ECC/MCE, "
                     "NIC state/errors, Bond slaves/failures, SR-IOV VF state, "
                     "Kernel critical messages, Systemd failed services, "
                     "NTP sync/offset, PCIe/AER errors, BIOS/NIC firmware, "
                     "NUMA topology, Hugepages, SELinux, Firewall, Listening ports"),
]

for row_idx, (label, value) in enumerate(instructions, start=2):
    is_section = label and not label.startswith(" ") and not label[0].isdigit() and not value and "(" not in label
    if is_section and label:
        ws3.merge_cells(f"A{row_idx}:E{row_idx}")
        c = ws3[f"A{row_idx}"]
        c.value = label
        c.font = Font(name="Consolas", bold=True, size=9, color=HDR_FG)
        c.fill = PatternFill("solid", fgColor=HDR_BG)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws3.row_dimensions[row_idx].height = 20
    elif label and value:
        c1 = ws3[f"A{row_idx}"]
        c1.value = label
        c1.font = Font(name="Consolas", size=9, color="60A5FA", bold=True)
        c1.fill = PatternFill("solid", fgColor="171923")
        c1.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False, indent=2)
        ws3.row_dimensions[row_idx].height = 16
        c2 = ws3[f"B{row_idx}"]
        c2.value = value
        c2.font = Font(name="Calibri", size=9, color="C8D0E0")
        c2.fill = PatternFill("solid", fgColor="171923")
        c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        ws3.merge_cells(f"B{row_idx}:E{row_idx}")
    else:
        ws3.row_dimensions[row_idx].height = 8

ws3.column_dimensions["A"].width = 40
ws3.column_dimensions["B"].width = 80
ws3.column_dimensions["C"].width = 10
ws3.column_dimensions["D"].width = 10
ws3.column_dimensions["E"].width = 10

# ── Tab order ─────────────────────────────────────────────────────────────────
wb.active = ws1

wb.save("/home/claude/clusterpulse/config/inventory.xlsx")
print("Inventory created OK")